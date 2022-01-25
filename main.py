
from asyncio.windows_events import NULL
from distutils.log import Log
from glob import glob
from msilib.schema import File
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import time
import enum
import eel
from openpyxl import load_workbook
import requests
import os
import shutil



#GLOBALS 
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
Armed = False
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER
#ARMING !!!! WILL MAKE IT WORK !!!!!!#DANGER DANGER DANGER DANGER

#is the program logged in? 
LoggedIn = False
#Whats the user name 
Username =""
#whats the password 
Password = ""
#LastName 
LastName = ""
#First Name
FirstName = ""
#open chrome 
driver = webdriver.Chrome()

CurrentVersion = 1.1
VersionIn = 0
updatelogs = []
DisplayedUpdate = False 
ChemsjustCleared = False
UpdateIssues = []
class RecCenterO(enum.Enum):   
    none = "none"
    Carla = "Carla"
    Central = "Central"

def LogError(errin):
    print("ERROR - " + errin)

def LogPrint(login):
    print("LOG - " + login)

CurRecCenter = RecCenterO.none



def ClearPause():
    time.sleep(.5)
    eel.ClearPause()
def ClearLoad():
    time.sleep(.5)
    eel.ClearLoad()


def CheckArm(): 
    if Armed == True: 
        eel.AddArmedWarn()


#close Everything 
@eel.expose
def EndProgram():
    LogPrint("EndProgram() Program Ending :(")
    driver.quit()
    quit(1)
 

#saves login to file
@eel.expose
def SetRecCenter(RecCenterIn):
    global CurRecCenter
    CurRecCenter = RecCenterO(RecCenterIn)
    LogPrint("SetRecCenter() Rec Center is now: " + str(RecCenterIn))
    
@eel.expose
def SetLoggedInFalse():
    global LoggedIn
    LoggedIn = False
@eel.expose
def SaveLogin(UsernameIn,PasswordIn, LastNameIn, FirstNameIn):
    LogPrint("SaveLogin() Starting Login Save")
    FiletoSave = open("input.txt", "w")
    TCurRecCenter = str(CurRecCenter.value)
    FiletoSave.write(UsernameIn.removesuffix("\n") +"\n")
    
    FiletoSave.write(PasswordIn.removesuffix("\n") + "\n")
    FiletoSave.write(TCurRecCenter.removesuffix("\n") + "\n")
    FiletoSave.write(LastNameIn.removesuffix("\n") + "\n")
    FiletoSave.write(FirstNameIn.removesuffix("\n") + "\n")
    FiletoSave.close()
    LogPrint("SaveLogin() Finished Login save")
@eel.expose
def SaveLoginA():
    SaveLogin(Username.removesuffix("\n"),Password.removesuffix("\n"), LastName.removesuffix("\n") , FirstName.removesuffix("\n"))

#logins to program 
def Login():
    LogPrint("Login() Attempting to login")
    global Username
    global Password
    global LoggedIn
    driver.get("https://www.digiquatics.com/patron_counts/new?location_id=10991")
    driver.find_element_by_id("user_email").send_keys(Username)
    driver.find_element_by_id("user_password").send_keys(Password)
    #driver.find_element_by_name("commit").click()
    LoggedIn = True
    #driver.get("https://www.digiquatics.com/patron_counts/new?location_id=10991")
    LogPrint("Login() Logged in!")

#get the size of a file 
@eel.expose
def GetFileSize(FileInThing):
    size = 0
    for line in FileInThing:
        size+=1
    return size


#Loads Login from file 
def LoadLogin():
    LogPrint("LoadLogin() starting Loading Login")
    #define usable globals 
    global Username
    global Password
    global LoggedIn
    global CurRecCenter
    global LastName
    global FirstName
    #open file "input.txt" for reading 
    FileInputIn = open("input.txt", "r")
    #check file size and setup size 
    
    size = GetFileSize(FileInputIn)
    #if there are the correct number of vars store them in correct vars 
    if size == 4 or size ==5:
        tx = 0
        FileInputIn = open("input.txt", "r")
        for line in FileInputIn:
            if tx == 0:
                Username = line
            if tx == 1:
                Password = line
            if tx ==2:
                if line == NULL or line == "" or line == "none":
                    LogError("Input file doesnt contain everything it need to")
                else:
                    CurRecCenter = RecCenterO(line.removesuffix("\n"))
                    eel.LoadRecCenter(str(line.removesuffix("\n")))
            if tx == 3:
                LastName = line
            if tx == 4: 
                FirstName = line
            tx+=1
        LogPrint("LoadLogin() Finished Load Login Sucessfully")
        return(True)
    else:
        #if there is not enough vars return false 
        LogError("LoadLogin() Login failed ")
        return(False)
        

#print from js 
@eel.expose
def PyPrint(print):
    print(print)

#the main function !
@eel.expose
def StartPro():
    LogPrint("StartPro() Starting Porgram")
    BLoadLogin = LoadLogin()
    #check if login was able to be done if not load the setup page if not login! 
    if BLoadLogin == False:
        eel.LoadSetupPage()
    elif BLoadLogin == True:
        if LoggedIn == False:
            Login()
    LogPrint("StartPro() Program Started ! ")

#enum for weekdays
class Day(enum.Enum):
    Monday = 1
    Tuesday = 2
    Wednesday = 3
    Thursday = 4
    Friday = 5
    Saturday= 6
    Sunday = 7

#ONE TIME takes DateIn(1/10/2022) DayofWeek(int = day of week 0 being monday ) LapIn(an array of all lap data) ActIn( an array of all Activitiy data )
@eel.expose
def PatronCheck(DateIn, DayofWeek, LapIn, ActIn):
    LogPrint("PatronCheck() Starting Patron Check Input ")
    #loads Patron Page 
    driver.get("https://www.digiquatics.com/patron_counts/new?location_id=10991")
    #Sets Wday up Monday is a placeholder
    Wday = Day.Monday
    #setup vars from js 
    #string of the raw date from JS 
    Date = DateIn
    #Array of the LAp 
    Lap = LapIn
    #array of the activity pool 
    Activity = ActIn

    #Based on the day Setup the ENUM 
    if DayofWeek == 0:
        Wday = Day.Monday
    if DayofWeek == 1:
        Wday = Day.Tuesday
    if DayofWeek == 2:
        Wday = Day.Wednesday
    if DayofWeek == 3:
        Wday = Day.Thursday
    if DayofWeek == 4:
        Wday = Day.Friday
    if DayofWeek == 5:
        Wday = Day.Saturday
    if DayofWeek == 6:
        Wday = Day.Sunday


    #tnum defind the hour of the day starting at 0 so if we open at 8 then 0 would be 8am 
    tnum = 0
    #number of hours in that speciifc day
    numhours = 0
    #start time of that day 
    starttime =0
    #where we are in the activity array  
    ActCount = 0
    #where we are int the lap array 
    LapCount = 0
    #detects day and prints any errors 
    if Wday == Day.Monday:
        numhours = 16
        starttime = 5
        if len(Activity) != 32:
            quit(2)
        if len(Lap) != 26:
            quit(2)
    if Wday == Day.Tuesday:
        numhours = 16
        starttime = 5
        if len(Activity) != 32:
            quit(2)
        if len(Lap) != 26:
            quit(2)
    if Wday == Day.Wednesday:
        numhours = 16
        starttime = 5
        if len(Activity) != 32:
            quit(2)
        if len(Lap) != 26:
            quit(2)
    if Wday == Day.Thursday:
        numhours = 16
        starttime = 5
        if len(Activity) != 32:
            quit(2)
        if len(Lap) != 26:
            quit(2)
    if Wday == Day.Friday:
        numhours = 15
        starttime = 5
        if len(Activity) != 28:
            quit(2)
        if len(Lap) != 21:
            quit(2)
    if Wday == Day.Saturday:
        numhours = 9
        starttime = 8
        if len(Activity) != 17:
            quit(2)
        if len(Lap) != 15:
            quit(2)
    if Wday == Day.Sunday:
        numhours = 9
        starttime = 8
        if len(Activity) != 17:
            quit(2)
        if len(Lap) != 15:
            quit(2)
    #FOR EACH DAY !!!
    for x in range(1,numhours+1):
        #click Date field
        driver.find_element_by_id("createdAtPicker").click()

        #clears date field with ctrl + a then del
        ActionChains(driver) \
            .key_down(Keys.CONTROL) \
            .key_down("a") \
            .key_up(Keys.CONTROL) \
            .key_up("a") \
            .key_down(Keys.DELETE) \
            .key_up(Keys.DELETE) \
            .perform()

        #Date = "11/13/2021 " #Month Day Year (leave space after !!!!) 
        #creates the Date from a hogpoge 
        DateTime = ((starttime-1)+x)%12
        texttime =""
        texttime += Date
        texttime += str(DateTime)
        texttime += ":00 "
        #determinds am or pm 
        if ((starttime-1)+x) >= 12:
            texttime += "PM"
        else:
            texttime += "AM"
        #sends keys and presses enter 
        driver.find_element_by_id("createdAtPicker").send_keys(texttime)
        driver.find_element_by_id("createdAtPicker").send_keys(Keys.ENTER)
        #driver.find_element_by_class_name("form-control").click()
        #needs sleep time for webpage to catch up with porgram (Py is fast) 
        time.sleep(.5)
        elementstest = driver.find_elements_by_class_name("form-control")



        # code Days
        #define inputs for each day 
        if Wday == Day.Monday:
            if tnum <=3:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])#Adult/River
                ActCount += 1
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys("0")#Swim Lessons
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys("0")#Spectators
                elementstest[8].send_keys("0")#Aqua Aerobics
            if tnum == 4:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])#Adult/River
                ActCount += 1
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys("0")#Swim Lessons
                elementstest[6].send_keys(Activity[ActCount])#Aqua Zumba
                ActCount += 1
                elementstest[7].send_keys(Activity[ActCount])#Spectators
                ActCount += 1
                elementstest[8].send_keys("0")#Aqua Aerobics
            if tnum == 5:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys("0")#Adult/River
                elementstest[6].send_keys(Activity[ActCount])#Aqua Zumba
                ActCount += 1
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")#Swim Lessons

                elementstest[7].send_keys(Activity[ActCount])#Spectators
                ActCount += 1
                elementstest[8].send_keys("0")#Aqua Aerobics

            if tnum >5 and tnum <9:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")#Swim Lessons
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys(Activity[ActCount])#Spectators
                ActCount += 1
                elementstest[8].send_keys("0")#Aqua Aerobics

            if tnum == 9 or tnum == 10 or tnum == 15:
                elementstest[1].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[2].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")#Swim Lessons
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1
                elementstest[8].send_keys("0")#Aqua Aerobics
            if tnum == 11:
                elementstest[1].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[2].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Swim Lessons
                LapCount += 1
                ActCount += 1
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1
                elementstest[8].send_keys("0")#Aqua Aerobics
            if tnum == 12:
                elementstest[1].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Swim Lessons
                LapCount += 1
                ActCount += 1
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1
                elementstest[8].send_keys("0")#Aqua Aerobics
            if tnum == 13:
                elementstest[2].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[1].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Swim Lessons
                LapCount += 1
                ActCount += 1
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1
                elementstest[8].send_keys("0")#Aqua Aerobics
            if tnum == 14:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")#Swim Lessons
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1
                elementstest[8].send_keys("0")#Aqua Aerobics
        if Wday == Day.Tuesday:
            if tnum <=3:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])#Adult/River
                ActCount += 1
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys("0")#Swim Lessons
                elementstest[6].send_keys("0")#Spectators

            if tnum == 4:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])#Adult/River
                ActCount += 1
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys(Activity[ActCount])#Swim Lessons
                ActCount += 1
                elementstest[6].send_keys(Activity[ActCount])#Spectators
                ActCount += 1

            if tnum == 5:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys("0")#Adult/River
                elementstest[5].send_keys(Activity[ActCount])#Swim Lessons
                ActCount += 1
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1


                elementstest[6].send_keys(Activity[ActCount])#Spectators
                ActCount += 1


            if tnum >5 and tnum <9:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")#Swim Lessons

                elementstest[6].send_keys(Activity[ActCount])#Spectators
                ActCount += 1


            if tnum == 9 or tnum == 10 or tnum == 15:
                elementstest[1].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[2].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")#Swim Lessons

                elementstest[6].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1

            if tnum == 11:
                elementstest[1].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[2].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Swim Lessons
                LapCount += 1
                ActCount += 1

                elementstest[6].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1

            if tnum == 12:
                elementstest[1].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Swim Lessons
                LapCount += 1
                ActCount += 1

                elementstest[6].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1

            if tnum == 13:
                elementstest[2].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[1].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Swim Lessons
                LapCount += 1
                ActCount += 1

                elementstest[6].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1

            if tnum == 14:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")#Swim Lessons

                elementstest[6].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1
        if Wday == Day.Wednesday:
            if tnum <=3:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])#Adult/River
                ActCount += 1
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys("0")#Swim Lessons
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys("0")#Spectators

            if tnum == 4:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])#Adult/River
                ActCount += 1
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys("0")#Swim Lessons
                elementstest[6].send_keys(Activity[ActCount])#Aqua Zumba
                ActCount += 1
                elementstest[7].send_keys(Activity[ActCount])#Spectators
                ActCount += 1

            if tnum == 5:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys("0")#Adult/River
                elementstest[6].send_keys(Activity[ActCount])#Aqua Zumba
                ActCount += 1
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")#Swim Lessons

                elementstest[7].send_keys(Activity[ActCount])#Spectators
                ActCount += 1


            if tnum >5 and tnum <9:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")#Swim Lessons
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys(Activity[ActCount])#Spectators
                ActCount += 1


            if tnum == 9 or tnum == 10 or tnum == 15:
                elementstest[1].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[2].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")#Swim Lessons
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1

            if tnum == 11:
                elementstest[1].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[2].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Swim Lessons
                LapCount += 1
                ActCount += 1
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1

            if tnum == 12:
                elementstest[1].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Swim Lessons
                LapCount += 1
                ActCount += 1
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1

            if tnum == 13:
                elementstest[2].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[1].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Swim Lessons
                LapCount += 1
                ActCount += 1
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1

            if tnum == 14:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")#Adult/River
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")#Swim Lessons
                elementstest[6].send_keys("0")#Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1
        if Wday == Day.Thursday:
            if tnum <= 3:
                elementstest[1].send_keys(Lap[LapCount])  # Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")  # Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])  # Adult/River
                ActCount += 1
                elementstest[4].send_keys("0")  # Open Swim (Leisure Pool)
                elementstest[5].send_keys("0")  # Swim Lessons
                elementstest[6].send_keys("0")  # Aqua Zumba
                elementstest[7].send_keys("0")  # Spectators

            if tnum == 4:
                elementstest[1].send_keys(Lap[LapCount])  # Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")  # Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])  # Adult/River
                ActCount += 1
                elementstest[4].send_keys("0")  # Open Swim (Leisure Pool)
                elementstest[5].send_keys("0")  # Swim Lessons
                elementstest[6].send_keys(Activity[ActCount])  # Aqua Zumba
                ActCount += 1
                elementstest[7].send_keys(Activity[ActCount])  # Spectators
                ActCount += 1

            if tnum == 5:
                elementstest[1].send_keys(Lap[LapCount])  # Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")  # Open Swim (Lap Pool)
                elementstest[3].send_keys("0")  # Adult/River
                elementstest[6].send_keys(Activity[ActCount])  # Aqua Zumba
                ActCount += 1
                elementstest[4].send_keys(Activity[ActCount])  # Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")  # Swim Lessons

                elementstest[7].send_keys(Activity[ActCount])  # Spectators
                ActCount += 1

            if tnum > 5 and tnum < 9:
                elementstest[1].send_keys(Lap[LapCount])  # Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")  # Open Swim (Lap Pool)
                elementstest[3].send_keys("0")  # Adult/River
                elementstest[4].send_keys(Activity[ActCount])  # Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")  # Swim Lessons
                elementstest[6].send_keys("0")  # Aqua Zumba
                elementstest[7].send_keys(Activity[ActCount])  # Spectators
                ActCount += 1

            if tnum == 9 or tnum == 10 or tnum == 15:
                elementstest[1].send_keys("0")  # Lap Swim (Lap Pool)
                elementstest[2].send_keys(Lap[LapCount])  # Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")  # Adult/River
                elementstest[4].send_keys(Activity[ActCount])  # Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")  # Swim Lessons
                elementstest[6].send_keys("0")  # Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))  # Spectators
                LapCount += 1
                ActCount += 1

            if tnum == 11:
                elementstest[1].send_keys("0")  # Lap Swim (Lap Pool)
                elementstest[2].send_keys(Lap[LapCount])  # Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")  # Adult/River
                elementstest[4].send_keys(Activity[ActCount])  # Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))  # Swim Lessons
                LapCount += 1
                ActCount += 1
                elementstest[6].send_keys("0")  # Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))  # Spectators
                LapCount += 1
                ActCount += 1

            if tnum == 12:
                elementstest[1].send_keys("0")  # Lap Swim (Lap Pool)
                elementstest[2].send_keys("0")  # Open Swim (Lap Pool)
                elementstest[3].send_keys("0")  # Adult/River
                elementstest[4].send_keys("0")  # Open Swim (Leisure Pool)
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))  # Swim Lessons
                LapCount += 1
                ActCount += 1
                elementstest[6].send_keys("0")  # Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))  # Spectators
                LapCount += 1
                ActCount += 1

            if tnum == 13:
                elementstest[2].send_keys("0")  # Lap Swim (Lap Pool)
                elementstest[1].send_keys(Lap[LapCount])  # Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")  # Adult/River
                elementstest[4].send_keys(Activity[ActCount])  # Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))  # Swim Lessons
                LapCount += 1
                ActCount += 1
                elementstest[6].send_keys("0")  # Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))  # Spectators
                LapCount += 1
                ActCount += 1

            if tnum == 14:
                elementstest[1].send_keys(Lap[LapCount])  # Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys(Lap[LapCount])  # Open Swim (Lap Pool)
                LapCount += 1
                elementstest[3].send_keys("0")  # Adult/River
                elementstest[4].send_keys(Activity[ActCount])  # Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys("0")  # Swim Lessons
                elementstest[6].send_keys("0")  # Aqua Zumba
                elementstest[7].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))  # Spectators
                LapCount += 1
                ActCount += 1
        if Wday == Day.Friday:
            if tnum <= 3:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])#Adult/River
                ActCount += 1
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys("0")#Aqua Zumba
                elementstest[6].send_keys("0")#Spectators
            if tnum == 4:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])#Adult/River
                ActCount += 1
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys(Activity[ActCount])#Aqua Zumba
                ActCount += 1
                elementstest[6].send_keys(Activity[ActCount])#Spectators
                ActCount += 1
            if tnum == 5:
                elementstest[1].send_keys(Lap[LapCount])  # Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")  # Open Swim (Lap Pool)
                elementstest[3].send_keys("0")  # Adult/River
                elementstest[5].send_keys(Activity[ActCount])  # Aqua Zumba
                ActCount += 1
                elementstest[4].send_keys(Activity[ActCount])  # Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[6].send_keys(Activity[ActCount])  # Spectators
                ActCount += 1
            if tnum > 5 and tnum < 9:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])#Adult/River
                ActCount += 1
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys("0")#Aqua Zumba
                elementstest[6].send_keys(Activity[ActCount])#Spectators
                ActCount += 1
            if tnum > 8:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])#Adult/River
                ActCount += 1
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys("0")#Aqua Zumba
                elementstest[6].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                ActCount += 1
                LapCount += 1
        if Wday == Day.Saturday:
            if tnum == 0 or tnum ==1:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount +=1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])#Adult/River
                ActCount+=1
                elementstest[4].send_keys("0")#Open Swim (Leisure Pool)
                elementstest[5].send_keys("0")#Spectators
            if tnum == 2:
                elementstest[1].send_keys(Lap[LapCount])#Lap Swim (Lap Pool)
                LapCount += 2
                elementstest[3].send_keys(Activity[ActCount])#Adult/River
                ActCount += 1
                elementstest[2].send_keys("0")#Open Swim (Lap Pool)


                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(Activity[ActCount])#Spectators
                ActCount += 1
            if tnum == 3:
                elementstest[1].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[3].send_keys("0")#Adult/River
                elementstest[2].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(Activity[ActCount])#Spectators
                ActCount += 1
            if tnum > 3:
                elementstest[1].send_keys("0")#Lap Swim (Lap Pool)
                elementstest[3].send_keys("0")#Adult/River
                elementstest[2].send_keys(Lap[LapCount])#Open Swim (Lap Pool)
                LapCount += 1
                elementstest[4].send_keys(Activity[ActCount])#Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))#Spectators
                LapCount += 1
                ActCount += 1
        if Wday == Day.Sunday:
            if tnum == 0 or tnum == 1:
                elementstest[1].send_keys(Lap[LapCount])  # Lap Swim (Lap Pool)
                LapCount += 1
                elementstest[2].send_keys("0")  # Open Swim (Lap Pool)
                elementstest[3].send_keys(Activity[ActCount])  # Adult/River
                ActCount += 1
                elementstest[4].send_keys("0")  # Open Swim (Leisure Pool)
                elementstest[5].send_keys("0")  # Spectators
            if tnum == 2:
                elementstest[1].send_keys(Lap[LapCount])  # Lap Swim (Lap Pool)
                LapCount += 2
                elementstest[3].send_keys(Activity[ActCount])  # Adult/River
                ActCount += 1
                elementstest[2].send_keys("0")  # Open Swim (Lap Pool)

                elementstest[4].send_keys(Activity[ActCount])  # Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(Activity[ActCount])  # Spectators
                ActCount += 1
            if tnum == 3:
                elementstest[1].send_keys("0")  # Lap Swim (Lap Pool)
                elementstest[3].send_keys("0")  # Adult/River
                elementstest[2].send_keys(Lap[LapCount])  # Open Swim (Lap Pool)
                LapCount += 1
                elementstest[4].send_keys(Activity[ActCount])  # Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(Activity[ActCount])  # Spectators
                ActCount += 1
            if tnum > 3:
                elementstest[1].send_keys("0")  # Lap Swim (Lap Pool)
                elementstest[3].send_keys("0")  # Adult/River
                elementstest[2].send_keys(Lap[LapCount])  # Open Swim (Lap Pool)
                LapCount += 1
                elementstest[4].send_keys(Activity[ActCount])  # Open Swim (Leisure Pool)
                ActCount += 1
                elementstest[5].send_keys(str(int(Lap[LapCount]) + int(Activity[ActCount])))  # Spectators
                LapCount += 1
                ActCount += 1

        tnum+=1
        #my Test Stuff !! 
        #input("COntinue????")






        #DANGER
        #DANGER
        #DANGER
        #DANGER
        #DANGER
        #DANGER
        if Armed == True:
            driver.find_element_by_class_name("btn-pink").click()
        #DANGER
        #DANGER
        #DANGER
        #DANGER
        #DANGER



        driver.get("https://www.digiquatics.com/patron_counts/new?location_id=10991")

    ClearPause()
    eel.AddDisplay("Finished!", "Patron Count ran successfully and finished execution")
    LogPrint("PatronCheck() Patron check sucessfull ")
CurWriteLineLap = 4 
CurWriteLineAct = 4 
@eel.expose
def UploadExcel():
    LogPrint("UploadExcel() Starting to upload EXCEL")
    driver.get("https://www.digiquatics.com/chemical_records/importer")
    FileUpload = driver.find_element_by_id("file")
    
    ExcelAdress = os.getcwd()
    ExcelAdress += "/ChemicalRecordTemplate.xlsx"
    NewExcelAdress = os.getcwd()
    NewExcelAdress += "/SubmittedExcel/Save.xlsx"
    shutil.copy2(ExcelAdress, NewExcelAdress); 

    driver.find_element_by_id("file").send_keys(ExcelAdress)
    if Armed == True:
        driver.find_element_by_class_name("btn-pink").click()
    ClearXcel()
    ClearPause()
    eel.AddDisplay("Uploaded!","sucessfully uploaded to digiquatics")
    LogPrint("UploadExcel() Sucessfully Uploaded Excel")

@eel.expose
def ChemCheck(CLIN, ORPIN, PHIN,TAIN,FLOWIN,TEMPIN,NOTESIN, DateIN, DayofWeekIN,LapOAct):
    LogPrint("ChemCheck() Starting Chem check ")
    DateIN.replace("-","/")
    global CurWriteLineLap
    global CurWriteLineAct
    AllTimes = ["Open", "10:30am", "02:30pm", "04:30pm","Close"]
    if DayofWeekIN == 5 or DayofWeekIN == 6: 
        AllTimes[2] = "12:30pm"
        AllTimes[3] = "02:30pm"
        AllTimes[4] = "04:30pm"
        AllTimes[0] = "08:00am"
    elif DayofWeekIN == 4:
        AllTimes[0] = "05:00am"
        AllTimes[4] = "07:00pm"
    else:
        AllTimes[0] = "05:00am"
        AllTimes[4] = "07:00pm"
    xcelname = "ChemicalRecordTemplate.xlsx"
    wb = load_workbook(filename= xcelname)
    LapSheet = wb["Central Park Central P...-9181"]
    ActSheet = wb["Central Park Central P...-9191"]
    if LapOAct == 0:
        x = 0
        while x < 5:
            LapSheet["A" + str(CurWriteLineLap)] = DateIN
            LapSheet["B" + str(CurWriteLineLap)] = AllTimes[x]
            LapSheet["C" + str(CurWriteLineLap)] = LastName + "," + FirstName
            LapSheet["F" + str(CurWriteLineLap)] = CLIN[x]
            LapSheet["G" + str(CurWriteLineLap)] = PHIN[x]
            LapSheet["F" + str(CurWriteLineLap)] = CLIN[x]
            LapSheet["J" + str(CurWriteLineLap)] = "Clear"
            if x == 0:
                LapSheet["K" + str(CurWriteLineLap)] = TAIN
            LapSheet["H" + str(CurWriteLineLap)] = TEMPIN[x]
            LapSheet["M" + str(CurWriteLineLap)] = ORPIN[x]
            LapSheet["N" + str(CurWriteLineLap)] = NOTESIN[x]
            x+=1 
            CurWriteLineLap += 1
    elif LapOAct == 1:
        x = 0
        while x < 5:
            ActSheet["A" + str(CurWriteLineAct)] = DateIN
            ActSheet["B" + str(CurWriteLineAct)] = AllTimes[x]
            ActSheet["C" + str(CurWriteLineAct)] = LastName + "," + FirstName
            ActSheet["F" + str(CurWriteLineAct)] = CLIN[x]
            ActSheet["G" + str(CurWriteLineAct)] = PHIN[x]
            ActSheet["F" + str(CurWriteLineAct)] = CLIN[x]
            ActSheet["J" + str(CurWriteLineAct)] = "Clear"
            if x == 0:
                ActSheet["K" + str(CurWriteLineAct)] = TAIN
            ActSheet["H" + str(CurWriteLineAct)] = TEMPIN[x]
            ActSheet["M" + str(CurWriteLineAct)] = ORPIN[x]
            ActSheet["N" + str(CurWriteLineAct)] = NOTESIN[x]

            x+=1 
            CurWriteLineAct += 1


    wb.save(xcelname)
    LogPrint("CheckChem() Sucessfully did Chem Check ")
    ClearPause() 
    
#Sets up Chem page 
@eel.expose
def SetupChem():
    CheckArm()
    LogPrint("SetupChem() Starting to setup chem ")
    global ChemsjustCleared
    if ChemsjustCleared == True: 
        eel.AddDisplay("Chems Cleared", "Sucessfully Cleared all chems")
        ChemsjustCleared = False 
    ClearLoad()
    LogPrint("SetupChem( Finished Setting up chem page ")
    #ClearXcel(); 
#Sets up Patron page 
@eel.expose
def SetupPatron():
    CheckArm()
    LogPrint("SetupPatron() Starting Setting up Patron")
    ClearLoad()
    LogPrint("SetupPatron() Finished setting up patron ")
@eel.expose
def SetupSetup():
    LogPrint("SetupSetup() Starting setting up Setup Page ")
    ClearLoad()
    LogPrint("SetupSetup() Finished setting up setup page ")

#Sets up Home page 
@eel.expose
def SetupHome():
    CheckArm()
    LogPrint("SetupHome() Started setting up Home Page ")
    #globals 
    global VersionIn
    global DisplayedUpdate
    #get pastbin 
    url = 'https://pastebin.com/raw/bB4wBAPP'
    req = requests.get(url)
    #open Version save file 
    FileVersion = open("Version.txt" , "r")
    #parse through pastebin
    txin = 0
    for line in FileVersion:
        if(txin == 0):
            CurrentVersion = float(line)
        txin += 1 
    #save File (not really should only be read)
    FileVersion.close()
    #Split into Lines
    Lines = req.text.split("\n")
    #Setup Vars
    cur = 0 
    for x in Lines: 
        if cur == 0: 
            VersionIn = float(x)
        elif DisplayedUpdate == False:
            updatelogs.append(str(x))
        cur +=1
    # if the version is less than what it is 
    if CurrentVersion < VersionIn:
        #add the text on Home Page
        eel.AddModal(VersionIn,updatelogs)
        DisplayedUpdate = True
        
    issuesurl = "https://pastebin.com/raw/DuzJekba"
    reqissues = requests.get(issuesurl)
    reqissuestxt = reqissues.text
    reqissuestxt = reqissuestxt.replace("\r",'')
    global UpdateIssues
    AllUpdateIssues = reqissuestxt.split("/New/")

    for x in AllUpdateIssues:
        currentissueline = 0
        SingleUpdateIssue = x.split("\n")
        SingleUpdateIssue.remove('')
        for y in SingleUpdateIssue:
            if currentissueline == 0:
                if float(y) == CurrentVersion:
                    
                    UpdateIssues = SingleUpdateIssue

            currentissueline +=1 

    eel.Issues(UpdateIssues)


    
    #get Current Changes 
    Fileupdatelogs = open("update" + str(CurrentVersion) + ".txt")
    tcur =0 
    UpdateLogupdatelogs = []
    #Parase File update logs and set up vars 
    for y in Fileupdatelogs: 
        if tcur == 0: 
            UpdateLogVersionIn = float(y)
        else:
            
            UpdateLogupdatelogs.append(str(y))
            
        tcur +=1
    eel.SetUpdateLogs(UpdateLogVersionIn, UpdateLogupdatelogs); 
    
    StartPro()
    ClearLoad()
    LogPrint("SetupHome() Finished Settingup Home page ")
    
@eel.expose
def SetupSettings():
    LogPrint("SetupSettings() Setting up settings page ")
    eel.SetArmedCheck(Armed)
    ClearLoad()
    LogPrint("SetupSettings() Finihsed setting up settings page ")
#clear Xcel Sheet ! 

def ClearXcel():
    LogPrint("ClearXcel() Starting to clear xcel sheet ")
    global CurWriteLineLap
    global CurWriteLineAct
    xcelname = "ChemicalRecordTemplate.xlsx"
    wb = load_workbook(filename= xcelname)
    LapSheet = wb["Central Park Central P...-9181"]
    ActSheet = wb["Central Park Central P...-9191"]
    x =4
    while x < 104: 
        y =0
        while y < 14:
            LapSheet[chr(65+y)+str(x)] = ""
            ActSheet[chr(65+y)+str(x)] = ""
            y+=1
        x+=1
    wb.save(xcelname)
    CurWriteLineLap =4
    CurWriteLineAct= 4
    LogPrint("ClearXcel() Finished Clearing Excel ")

@eel.expose
def ReloadClearChem():
    LogPrint("ReloadClearChem() Starting to clear and reload chem")
    global ChemsjustCleared
    ClearXcel()
    ClearPause()
    eel.gotoChem()
    ChemsjustCleared = True
    LogPrint("ReloadClearChem() Finished clearing and reloading chem")
    
@eel.expose
def ReportBug():
    driver.get("https://github.com/harri665/PatronCountsDist/issues")
  

@eel.expose
def SetArm(BIN): 
    global Armed
    Armed = BIN
    LogPrint("Armed Changed to: " + str(BIN))
def main():
    #goes to folder Website and launcher main.html 
    
    eel.init('Website')
    eel.start('home.html')

#if its main run main 
if __name__ == "__main__":
    main()
